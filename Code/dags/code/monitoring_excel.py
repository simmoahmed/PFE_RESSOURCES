import os
import pandas as pd
from pyhive import hive
from openpyxl import load_workbook
from datetime import datetime
import glob
from unidecode import unidecode
import warnings
import numpy as np
warnings.filterwarnings("ignore")

class ExcelOps:
    @staticmethod
    def create_hive_database(database_name):
        conn = hive.Connection(host="host.docker.internal", port=10000)
        cursor = conn.cursor()
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {database_name}")
        conn.commit()
        conn.close()
        print(f"------------Database '{database_name}' created successfully------------")

    @staticmethod
    def create_or_update_hive_table(database_name, table_name, schema):
        conn = hive.Connection(host="host.docker.internal", port=10000)
        cursor = conn.cursor()
        cursor.execute(f"USE {database_name}")
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        existing_tables = cursor.fetchall()

        if existing_tables:
            print(f"------------Table '{table_name}' already exists in the database '{database_name}'------------")
            existing_schema = ExcelOps.get_hive_table_schema(database_name, table_name)

            # Add new columns if not exist
            for column, data_type in schema.items():
                normalized_column = column.replace(' ', '_').replace('é', 'e').replace('à', 'a').replace('ç', 'c').lower()
                if normalized_column not in (col.lower() for col in existing_schema):
                    try:
                        cursor.execute(f"ALTER TABLE {table_name} ADD COLUMNS (`{normalized_column}` {data_type})")
                    except Exception as e:
                        print(f"Error adding column {normalized_column}: {str(e)}")
            
            # Remove columns that no longer exist in the new schema
            for column in existing_schema.keys():
                if column not in schema:
                    try:
                        cursor.execute(f"ALTER TABLE {table_name} REPLACE COLUMNS ({', '.join([f'`{col}` {dtype}' for col, dtype in schema.items()])})")
                    except Exception as e:
                        print(f"Error removing column {column}: {str(e)}")

            conn.commit()
        else:
            create_table_query = f"CREATE TABLE IF NOT EXISTS {table_name} ("
            for column, data_type in schema.items():
                normalized_column = column.replace(' ', '_').replace('é', 'e').replace('à', 'a').replace('ç', 'c').replace('(', '').replace(')', '').lower()
                create_table_query += f"`{normalized_column}` {data_type}, "
            create_table_query = create_table_query[:-2]  
            create_table_query += f")"
            
            cursor.execute(create_table_query)
            conn.commit()
            print(f"------------Table '{table_name}' created successfully in database '{database_name}'------------")
        
        conn.close()

    @staticmethod
    def clean_value(value):
        if pd.isnull(value):
            return 'null'
        elif isinstance(value, str):
            cleaned_value = unidecode(value).replace('\n', ' ').replace('\r', ' ')
            cleaned_value = cleaned_value.replace("'", "")
            return f"'{cleaned_value}'"
        elif isinstance(value, datetime):
            return f"'{value}'"
        elif isinstance(value, (int, float)):
            return str(value)
        else:
            return "'"+str(value)+"'"

    @staticmethod
    def clean_valuee(value):
        if pd.isnull(value):
            return 'nan'
        elif isinstance(value, str):
            cleaned_value = unidecode(value).replace('\n', ' ').replace('\r', ' ')
            cleaned_value = cleaned_value.replace("'", "")
            return cleaned_value
        elif isinstance(value, datetime):
            return str(value)
        elif isinstance(value, (int, float)):
            return str(value)
        else:
            return str(value)
        
    @staticmethod
    def stringify_dataframe(df):
        return df.applymap(ExcelOps.clean_valuee)
    
    @staticmethod
    def apply_unidecode_to_dataframe(df):
        """Apply unidecode to all string columns in a DataFrame."""
        string_columns = df.select_dtypes(include=['object'])
        for col in string_columns.columns:
            df[col] = df[col].apply(lambda x: unidecode(x).replace("'", "") if isinstance(x, str) else x)
        return df

    @staticmethod
    def insert_data_into_hive_table(database_name, table_name, data):
        # Connect to Hive
        conn = hive.Connection(host="host.docker.internal", port=10000)
        cursor = conn.cursor()
        cursor.execute(f"USE {database_name}")
        
        # Clean and prepare data for insertion
        data = ExcelOps.apply_unidecode_to_dataframe(data)
        data = ExcelOps.stringify_dataframe(data)
        
        # Check if table exists and fetch existing data columns
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        existing_tables = cursor.fetchall()
        
        if existing_tables:
            # Delete existing data from the table (truncate)
            truncate_query = f"TRUNCATE TABLE {table_name}"
            cursor.execute(truncate_query)
            print(f"Truncated existing data from table '{table_name}'")
        
            # Insert new data into the table
            if not data.empty:
                data = data.where(pd.notnull(data), None)  # Replace NaN with None
                insert_query = f"INSERT INTO {table_name} VALUES "
                for index, row in data.iterrows():
                    values = ', '.join([str(ExcelOps.clean_value(value)) if value is not None else 'NULL' for value in row])
                    insert_query += f"({values}),"
                
                if insert_query.endswith(','):
                    insert_query = insert_query[:-1]  # Remove trailing comma
                
                cursor.execute(insert_query)
                print(f"Inserted {data.shape[0]} rows into table '{table_name}'")
        else:
            print(f"Table '{table_name}' does not exist in database '{database_name}'. Please create the table first.")
        
        # Commit changes and close connection
        conn.commit()
        conn.close()

    @staticmethod
    def get_hive_table_schema(database_name, table_name):
        conn = hive.Connection(host="host.docker.internal", port=10000)
        cursor = conn.cursor()
        cursor.execute(f"USE {database_name}")
        cursor.execute(f"DESCRIBE {table_name}")
        table_schema = {}
        for row in cursor.fetchall():
            column_name = row[0]
            data_type = row[1]
            table_schema[column_name] = data_type
        conn.close()
        return table_schema

    @staticmethod
    def infer_column_types(df):
        for col in df.columns:
            if df[col].dtype in ['int64', 'float64']:
                continue
            try:
                df[col] = pd.to_datetime(df[col], infer_datetime_format=True)
            except:
                continue
        column_types = {}
        for column in df.columns:
            for value in df[column]:
                if isinstance(value, int):
                    column_types[column] = 'BIGINT'
                    break
                elif isinstance(value, float):
                    column_types[column] = 'DOUBLE'
                    break
                elif isinstance(value, datetime):
                    column_types[column] = 'DATE'
                    df[column] = pd.to_datetime(df[column]).dt.date
                    break
                elif pd.api.types.is_string_dtype(df[column]):
                    column_types[column] = 'STRING'
                    break
        return column_types

def clean_excel_data(df):
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.replace('\n', ' ').str.replace('\r', ' ')
    return df

def excel_to_hive(excel_file):
    xls = load_workbook(excel_file, read_only=True, data_only=True)
    database_name = os.path.splitext(os.path.basename(excel_file))[0].replace('.', '').replace(' ', '_').replace('é', 'e').replace('à', 'a').replace('ç', 'c')
    ExcelOps.create_hive_database(database_name)
    for sheet_name in xls.sheetnames:
        sheet = xls[sheet_name]
        if not sheet.sheet_state == 'hidden':
            df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')
            df = clean_excel_data(df)
            column_types = ExcelOps.infer_column_types(df)
            table_name = sheet_name.replace('.', '').replace(' ', '_').replace('é', 'e').replace('à', 'a').replace('ç', 'c')
            ExcelOps.create_or_update_hive_table(database_name, table_name, column_types)
            ExcelOps.insert_data_into_hive_table(database_name, table_name, df)

def process_new_excel_files():
    path = "/opt/airflow/dags/data"
    excel_files = glob.glob(os.path.join(path, "*.xlsx"))
    for excel_file in excel_files:
        excel_to_hive(excel_file)

if __name__ == "__main__":
    process_new_excel_files()
